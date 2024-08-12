import os
from openpyxl import load_workbook

# Parse xlsx file to consts
root_dir = os.path.join(os.path.dirname(__file__), '..', '..')
xlsx_filename = os.path.join(root_dir, 'Designs', 'Constants.xlsx')
output_filename = os.path.join(root_dir, 'Scripts', 'Contents', 'Consts.cs')

# Load the workbook and select the active worksheet
workbook = load_workbook(xlsx_filename)
sheet = workbook.active

# Skip the header row and create a list of rows (excluding the header)
rows = list(sheet.iter_rows(min_row=2, values_only=True))

# Sort the rows by the first column (Name)
sorted_rows = sorted(rows, key=lambda x: x[3])

with open(output_filename, 'w') as outfile:
    outfile.write('public static class Consts\n')
    outfile.write('{\n')
    line_count = 1

    for row in sorted_rows:
        line_count += 1

        try:
            name, Type, Value, Category, description = row
        except:
            print(f"Error: Unable to process row {row} on line {line_count}")
            continue

        if type(Value) == str:
            # Escape single quotes in content, if any
            Value = Value.replace("'", "\\'")

        outfile.write(f'    public const {Type} {name} = {Value};  // {Category}: {description}\n')

    outfile.write('}\n')

print("Finished")
