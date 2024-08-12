import os
from openpyxl import load_workbook

# Parse xlsx file to enums
root_dir = os.path.join(os.path.dirname(__file__), '..', '..')
xlsx_filename = os.path.join(root_dir, 'Designs', 'Enums.xlsx')
output_filename = os.path.join(root_dir, 'Scripts', 'Contents', 'Enums.cs')

# Load the workbook and select the active worksheet
workbook = load_workbook(xlsx_filename)
sheet = workbook.active

# Skip the header row and create a list of rows (excluding the header)
rows = list(sheet.iter_rows(min_row=2, values_only=True))

# Sort the rows by the second column (category)
sorted_rows = sorted(rows, key=lambda x: x[1])

with open(output_filename, 'w') as outfile:
    outfile.write('public static class Enums\n')
    outfile.write('{\n')
    current_category = None
    line_count = 1

    for row in sorted_rows:
        line_count += 1

        try:
            name, category, description = row  # Adjust the indices if necessary
        except:
            print(f"Error: Unable to process row {row} on line {line_count}")
            continue

        if category != current_category:
            if current_category is not None:
                outfile.write('    }\n\n')  # Close the previous enum
            current_category = category
            outfile.write(f'    public enum {category}\n')  # Open a new enum
            outfile.write('    {\n')

        outfile.write(f'        {name},  // {description}\n')

    outfile.write('    }\n')  # Close the last enum
    outfile.write('}\n')

print("Finished")